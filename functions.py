import os
import xlrd
import pandas
import requests
import openpyxl
import statistics as st
from datetime import datetime
from dateutil.relativedelta import relativedelta

# Opening required spreadsheets for the code
wb = xlrd.open_workbook('custo_concorrentes.xlsx')
ws_concorrentes = wb.sheet_by_name('concorrentes')
ws_relacoes = wb.sheet_by_name('relacoes')
ws_multiplicadores = wb.sheet_by_name('multiplicadores')

# Access multipliers spreadsheet and getting all required questions
def getMultQuestions():
    questions = []
    j = 0
    for i in range(2, ws_multiplicadores.ncols, 3):
        questions.append(str(ws_multiplicadores.cell(1, i)))
        questions[j] = questions[j][6:-1]
        j += 1
    return questions

# Access relations spreadsheet and get both subservice and service
def getServiceRelations():
    # Index 0: subservice code
    # Index 1: service code
    service = [ws_relacoes.col_values(0), ws_relacoes.col_values(1)]
    service[0].pop(0)
    service[1].pop(0)
    return service

# Deletes current custo_concorrentes
# and make a request to download updated version of spreadsheet
def updateSpreadsheet():
    file_name = 'custo_concorrentes.xlsx'

    if os.path.isfile(file_name):
        os.remove(file_name)

    dls = 'https://trinuscapital-my.sharepoint.com/:x:/g/personal/arthur_pires_servmaisadm_com_br/ERkAAAEpiNhCpssP6o1-FQUBE2F2QVhJt23iN5eYMYT3HA?e=HOt2Pc&download=1'
    resp = requests.get(dls)

    output = open(file_name, 'wb')
    output.write(resp.content)
    output.close()

# Deletes current incc.xlsx, downloads another with fixed link (may cause future problems)
# Then, apply INCC to data in updated custo_concorrentes.xlsx
def updateINCC():
    file_name = 'incc.xlsx'

    if os.path.isfile(file_name):
        os.remove(file_name)

    dls = 'https://sindusconpr.com.br/download/8933/310'
    resp = requests.get(dls)

    output = open(file_name, 'wb')
    output.write(resp.content)
    output.close()

    # Accessing INCC's spreadsheet
    wb = xlrd.open_workbook('incc.xlsx')
    ws = wb.sheet_by_name('Plan1')

    # Getting INCC and clearing the data
    inccData = ws.col_values(1)
    for i in range(3):
        inccData.pop(0)
    while inccData[-1] == '':
        inccData.pop()

    # Creating list with dates from 1994 to last INCC date available
    date_format = '%d/%m/%Y'
    firstDate = datetime.strptime('01/08/1994', date_format)
    month = [firstDate]
    for i in range(len(inccData) - 1):
        month.append(month[i] + relativedelta(months=1))

    # Acessing concurrent's spreadsheet 
    wb = xlrd.open_workbook('custo_concorrentes.xlsx')
    ws = wb.sheet_by_name('concorrentes')

    # Getting data: columns K, L, M
    # K: custo unitario; L: custo total; M: custo_m2_lote
    concData = [ws.col_values(3), ws.col_values(12)]
    unCost = ws.col_values(10)
    totalCost = ws.col_values(11)

    # Cleaning data above
    unCost.pop(0)
    totalCost.pop(0)
    concData[0].pop(0)
    concData[1].pop(0)
    for i in range(len(concData[0])):
        concData[0][i] = xlrd.xldate_as_datetime(concData[0][i],0)

    # Apply INCC calculations to columns K, L, M
    for i in range(len(concData[0])):
        if concData[1][i] != -1:
            diff = relativedelta(concData[0][i], firstDate)
            delta = diff.years*12 + diff.months
            concData[1][i] = concData[1][i]/inccData[delta]*inccData[-1]
            unCost[i] = unCost[i]/inccData[delta]*inccData[-1]
            totalCost[i] = totalCost[i]/inccData[delta]*inccData[-1]

    file = openpyxl.load_workbook('custo_concorrentes.xlsx')
    sheet = file.get_sheet_by_name('concorrentes')
    for i in range(len(concData[1])):
        cell_avgCostArea = 'M' + str(i+2)
        cell_unitCost = 'K' + str(i+2)
        cell_totalCost = 'L' + str(i+2)
        sheet[cell_avgCostArea] = concData[1][i]
        sheet[cell_unitCost] = unCost[i]
        sheet[cell_totalCost] = totalCost[i]
    file.save('custo_concorrentes.xlsx')

def generateSpreadsheet(landArea, multipĺiers, serviceID, areaViario=None, paving='O8C3'):
    # Getting column values
    # G, J, L, M, N, O
    cod_subserv = ws_concorrentes.col_values(6)
    quantitativo = ws_concorrentes.col_values(9)
    totalCost = ws_concorrentes.col_values(11)
    custo_m2_lote = ws_concorrentes.col_values(12)
    area_total_lotes = ws_concorrentes.col_values(13)

    # Cleaning the data above
    cod_subserv.pop(0)
    quantitativo.pop(0)
    totalCost.pop(0)
    custo_m2_lote.pop(0)
    area_total_lotes.pop(0)

    ssDict = {}
    for i in range(len(cod_subserv)):
        # Removing useless data... (not even adding it)
        if custo_m2_lote[i] == -1: 
            continue
        
        # Finding the multiplier value for subservice
        relations = getServiceRelations()
        
        serviceList = ws_multiplicadores.col_values(0)
        for j in range(3): serviceList.pop(0)

        for z in range(len(relations[0])):
            if relations[0][z] == ws_concorrentes.cell(i+1, 6).value:
                service = relations[1][z]
                for j in range(len(serviceList)):
                    if serviceList[j] == service:
                        multiplier = multipĺiers[j]
                        break
                    else: multiplier = 1

        if ws_concorrentes.cell(i+1, 6).value not in ssDict.keys():
            ssDict[ws_concorrentes.cell(i+1, 6).value] = {
                                        'cod_subserv': ws_concorrentes.cell(i+1, 6).value,
                                        'descricao_subserv': ws_concorrentes.cell(i+1,7).value,
                                        'unidade': ws_concorrentes.cell(i+1,8).value,
                                        'quantitativo': [quantitativo[i]/area_total_lotes[i]*landArea],
                                        'qtd_m2': [quantitativo[i]/area_total_lotes[i]],
                                        'custo_un_insumo': [totalCost[i]/quantitativo[i]],
                                        'multiplicador': multiplier,
                                        'custo_calculado': 0,
                                        'custo_m2_lote': [custo_m2_lote[i]]                                 
                                    }
        else:
            ssDict[ws_concorrentes.cell(i+1, 6).value]['quantitativo'].append(quantitativo[i]/area_total_lotes[i]*landArea)
            ssDict[ws_concorrentes.cell(i+1, 6).value]['qtd_m2'].append(quantitativo[i]/area_total_lotes[i])
            ssDict[ws_concorrentes.cell(i+1, 6).value]['custo_un_insumo'].append(totalCost[i]/quantitativo[i])
            ssDict[ws_concorrentes.cell(i+1, 6).value]['custo_m2_lote'].append(custo_m2_lote[i])


    # Checking for the existence of the file and deleting it to create another.
    file_name = f'{serviceID}.xlsx'
    if os.path.isfile(file_name):
        os.remove(file_name)

    paving_options = ['O8C1', 'O8C2', 'O8C3']
    paving_options.remove(paving)
    del ssDict[paving_options[0]]
    del ssDict[paving_options[1]]

    for key in ssDict.keys():
        ssDict[key]['qtd_m2'] = round(st.mean(ssDict[key]['qtd_m2']), 8)
        ssDict[key]['quantitativo'] = round(st.mean(ssDict[key]['quantitativo']), 2)
        ssDict[key]['custo_un_insumo'] = round(st.mean(ssDict[key]['custo_un_insumo']), 2)
        ssDict[key]['custo_m2_lote'] = round(st.mean(ssDict[key]['custo_m2_lote']), 2)
        ssDict[key]['custo_calculado'] = ssDict[key]['quantitativo'] * ssDict[key]['custo_un_insumo'] * ssDict[key]['multiplicador']
        
    dataFrame = {
        'Sub-serviço': [],
        'Descrição': [],
        'Unidade': [],
        'Quantitativo': [],
        'Quantidade/m2': [],
        'Custo/unidade de insumo': [],
        'Multiplicadores': [],
        'Custo calculado': [],
    }
    for value in ssDict.values():
        dataFrame['Sub-serviço'].append(value['cod_subserv'])
        dataFrame['Descrição'].append(value['descricao_subserv'])
        dataFrame['Unidade'].append(value['unidade'])
        dataFrame['Quantitativo'].append(value['quantitativo'])
        dataFrame['Quantidade/m2'].append(value['qtd_m2'])
        dataFrame['Custo/unidade de insumo'].append(value['custo_un_insumo'])
        dataFrame['Multiplicadores'].append(value['multiplicador'])
        dataFrame['Custo calculado'].append(value['custo_calculado'])

    dataFrame = pandas.DataFrame(dataFrame)
    dataFrame.to_excel(file_name, sheet_name="Plan1", index=False)

def getMultipliers(answers):
    multipliers = []
    j = 0
    for i in range(2, ws_multiplicadores.ncols, 3):
        multipliers.append(ws_multiplicadores.col_values(i + answers[j]))
        for i in range(3):
            multipliers[j].pop(0)
        j += 1
    values = []
    for i in range(len(multipliers[0])):
        temp = 1
        for j in range(len(multipliers)):
            temp *= multipliers[j][i]
        values.append(temp)
    return values

def getCalculatedCost(serviceID):
    spreadsheetName = f'{serviceID}.xlsx'
    wb = xlrd.open_workbook(spreadsheetName)
    ws = wb.sheet_by_name('Plan1')

    values = ws.col_values(7)
    values.pop(0)
    calculatedCost = sum(values)
    return calculatedCost